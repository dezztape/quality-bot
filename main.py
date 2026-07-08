import asyncio
import os
import logging
from datetime import date, datetime, timedelta, timezone
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from dotenv import load_dotenv

from aiogram import Bot, Dispatcher, F
from aiogram.enums import ParseMode
from aiogram.exceptions import TelegramServerError
from aiogram.filters import Command
from aiogram.client.default import DefaultBotProperties

from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from aiogram.fsm.storage.memory import MemoryStorage

from aiogram.types import (
    Message,
    CallbackQuery,
    BotCommand,
    FSInputFile,
    ReplyKeyboardMarkup,
    KeyboardButton,
)

from aiogram.utils.keyboard import InlineKeyboardBuilder

from sqlalchemy import (
    BigInteger,
    Boolean,
    CheckConstraint,
    Date,
    DateTime,
    ForeignKey,
    Integer,
    String,
    UniqueConstraint,
    func,
    text,
    select,
)

from sqlalchemy.ext.asyncio import (
    AsyncSession,
    create_async_engine,
    async_sessionmaker,
)

from sqlalchemy.orm import DeclarativeBase, Mapped, mapped_column, relationship


# =========================
# ENV
# =========================

load_dotenv()

# Логирование
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)
logger.info("🤖 Quality Bot starting...")

BOT_TOKEN = os.getenv("BOT_TOKEN")
DATABASE_URL = os.getenv("DATABASE_URL")
REDIS_URL = os.getenv("REDIS_URL")
ADMIN_IDS = [
    int(x)
    for x in os.getenv("ADMIN_IDS", "").split(",")
    if x
]
MODERATOR_IDS = [
    int(x)
    for x in os.getenv("MODERATOR_IDS", "").split(",")
    if x
]
STATISTICIAN_IDS = [
    int(x)
    for x in os.getenv("STATISTICIAN_IDS", "").split(",")
    if x
]

logger.info(
    f"✅ Loaded {len(ADMIN_IDS)} admin(s), "
    f"{len(MODERATOR_IDS)} moderator(s), "
    f"{len(STATISTICIAN_IDS)} statistician(s)"
)


# =========================
# BOT
# =========================

bot = Bot(
    token=BOT_TOKEN,
    default=DefaultBotProperties(
        parse_mode=ParseMode.HTML
    )
)

dp = Dispatcher(storage=MemoryStorage())

if REDIS_URL:
    logger.info("📝 Redis URL configured (will use for production)")
else:
    logger.warning("⚠️ REDIS_URL not set, using MemoryStorage (development mode)")


# =========================
# DATABASE
# =========================

DB_SCHEMA = "quality_bot"
INIT_SCHEMA_SQL = Path(__file__).resolve().parent / "db" / "init_schema.sql"

# Преобразуем DATABASE_URL для asyncpg если нужно
# Railway отправляет postgresql://, нам нужен postgresql+asyncpg://
if DATABASE_URL and "postgresql+asyncpg" not in DATABASE_URL:
    DATABASE_URL = DATABASE_URL.replace("postgresql://", "postgresql+asyncpg://", 1)
    logger.info("🔄 Converted DATABASE_URL to use asyncpg")

engine = create_async_engine(
    DATABASE_URL,
    echo=False,
)

async_session = async_sessionmaker(
    engine,
    expire_on_commit=False,
)


class Base(DeclarativeBase):
    pass


# =========================
# MODELS
# =========================

class User(Base):
    __tablename__ = "users"
    __table_args__ = {"schema": DB_SCHEMA}

    id: Mapped[int] = mapped_column(BigInteger, primary_key=True)

    telegram_id: Mapped[int] = mapped_column(
        BigInteger,
        unique=True,
    )

    full_name: Mapped[str] = mapped_column(
        String(255)
    )

    city: Mapped[str] = mapped_column(
        String(255),
        default="",
    )

    is_whitelisted: Mapped[bool] = mapped_column(
        Boolean,
        default=False,
    )

    created_at: Mapped[datetime] = mapped_column(
        DateTime(timezone=True),
        server_default=func.now(),
    )

    updated_at: Mapped[datetime] = mapped_column(
        DateTime(timezone=True),
        server_default=func.now(),
        onupdate=func.now(),
    )

    sessions: Mapped[list["TestSession"]] = relationship(
        back_populates="user"
    )

    answers: Mapped[list["Answer"]] = relationship(
        back_populates="user"
    )


class Test(Base):
    __tablename__ = "tests"
    __table_args__ = (
        UniqueConstraint("test_date", name="uq_quality_bot_tests_test_date"),
        {"schema": DB_SCHEMA},
    )

    id: Mapped[int] = mapped_column(BigInteger, primary_key=True)

    title: Mapped[str] = mapped_column(
        String(255)
    )

    test_date: Mapped[date] = mapped_column(
        Date
    )

    active: Mapped[bool] = mapped_column(
        Boolean,
        default=True,
    )

    approved: Mapped[bool] = mapped_column(
        Boolean,
        default=False,
    )

    approved_at: Mapped[datetime | None] = mapped_column(
        DateTime(timezone=True),
        nullable=True,
    )

    created_at: Mapped[datetime] = mapped_column(
        DateTime(timezone=True),
        server_default=func.now(),
    )

    updated_at: Mapped[datetime] = mapped_column(
        DateTime(timezone=True),
        server_default=func.now(),
        onupdate=func.now(),
    )

    questions: Mapped[list["Question"]] = relationship(
        back_populates="test"
    )

    sessions: Mapped[list["TestSession"]] = relationship(
        back_populates="test"
    )


class Question(Base):
    __tablename__ = "questions"
    __table_args__ = (
        UniqueConstraint(
            "test_id",
            "test_type",
            "question_order",
            name="uq_quality_bot_questions_order",
        ),
        {"schema": DB_SCHEMA},
    )

    id: Mapped[int] = mapped_column(BigInteger, primary_key=True)

    test_id: Mapped[int] = mapped_column(
        BigInteger,
        ForeignKey(f"{DB_SCHEMA}.tests.id")
    )

    test_type: Mapped[str] = mapped_column(
        String(50),
        default="control",
    )

    question_order: Mapped[int] = mapped_column(
        Integer
    )

    image_file_id: Mapped[str] = mapped_column(
        String(500)
    )

    correct_answer: Mapped[int] = mapped_column(
        Integer
    )

    created_at: Mapped[datetime] = mapped_column(
        DateTime(timezone=True),
        server_default=func.now(),
    )

    updated_at: Mapped[datetime] = mapped_column(
        DateTime(timezone=True),
        server_default=func.now(),
        onupdate=func.now(),
    )

    test: Mapped["Test"] = relationship(
        back_populates="questions"
    )

    answers: Mapped[list["Answer"]] = relationship(
        back_populates="question"
    )


class Answer(Base):
    __tablename__ = "answers"
    __table_args__ = (
        UniqueConstraint(
            "attempt_id",
            "question_id",
            name="uq_quality_bot_answers_attempt_question",
        ),
        {"schema": DB_SCHEMA},
    )

    id: Mapped[int] = mapped_column(BigInteger, primary_key=True)

    attempt_id: Mapped[int] = mapped_column(
        BigInteger,
        ForeignKey(f"{DB_SCHEMA}.test_attempts.id")
    )

    user_id: Mapped[int] = mapped_column(
        BigInteger,
        ForeignKey(f"{DB_SCHEMA}.users.id")
    )

    question_id: Mapped[int] = mapped_column(
        BigInteger,
        ForeignKey(f"{DB_SCHEMA}.questions.id")
    )

    answer: Mapped[int] = mapped_column(
        Integer
    )

    is_correct: Mapped[bool] = mapped_column(
        Boolean
    )

    correct_answer_at_time: Mapped[int] = mapped_column(
        Integer
    )

    answered_at: Mapped[datetime] = mapped_column(
        DateTime(timezone=True),
        server_default=func.now(),
    )

    attempt: Mapped["TestAttempt"] = relationship(
        back_populates="answers"
    )

    user: Mapped["User"] = relationship(
        back_populates="answers"
    )

    question: Mapped["Question"] = relationship(
        back_populates="answers"
    )


class TestSession(Base):
    __tablename__ = "test_sessions"
    __table_args__ = (
        UniqueConstraint(
            "user_id",
            "test_id",
            "test_type",
            name="uq_quality_bot_test_sessions_user_test_type",
        ),
        {"schema": DB_SCHEMA},
    )

    id: Mapped[int] = mapped_column(BigInteger, primary_key=True)

    user_id: Mapped[int] = mapped_column(
        BigInteger,
        ForeignKey(f"{DB_SCHEMA}.users.id")
    )

    test_id: Mapped[int] = mapped_column(
        BigInteger,
        ForeignKey(f"{DB_SCHEMA}.tests.id")
    )

    test_type: Mapped[str] = mapped_column(
        String(50),
        default="control",
    )

    current_question: Mapped[int] = mapped_column(
        Integer,
        default=0,
    )

    completed: Mapped[bool] = mapped_column(
        Boolean,
        default=False,
    )

    started_at: Mapped[datetime] = mapped_column(
        DateTime(timezone=True),
        server_default=func.now(),
    )

    completed_at: Mapped[datetime | None] = mapped_column(
        DateTime(timezone=True),
        nullable=True,
    )

    attempts: Mapped[int] = mapped_column(
        Integer,
        default=1,
    )

    user: Mapped["User"] = relationship(
        back_populates="sessions"
    )

    test: Mapped["Test"] = relationship(
        back_populates="sessions"
    )

    attempts_history: Mapped[list["TestAttempt"]] = relationship(
        back_populates="session"
    )


class TestAttempt(Base):
    __tablename__ = "test_attempts"
    __table_args__ = (
        UniqueConstraint(
            "session_id",
            "attempt_number",
            name="uq_quality_bot_test_attempts_session_number",
        ),
        {"schema": DB_SCHEMA},
    )

    id: Mapped[int] = mapped_column(BigInteger, primary_key=True)

    session_id: Mapped[int] = mapped_column(
        BigInteger,
        ForeignKey(f"{DB_SCHEMA}.test_sessions.id")
    )

    attempt_number: Mapped[int] = mapped_column(
        Integer,
        default=1,
    )

    started_at: Mapped[datetime] = mapped_column(
        DateTime(timezone=True),
        server_default=func.now(),
    )

    completed_at: Mapped[datetime | None] = mapped_column(
        DateTime(timezone=True),
        nullable=True,
    )

    passed: Mapped[bool] = mapped_column(
        Boolean,
        default=False,
    )

    session: Mapped["TestSession"] = relationship(
        back_populates="attempts_history"
    )

    answers: Mapped[list["Answer"]] = relationship(
        back_populates="attempt"
    )


# =========================
# STATES
# =========================

class QuestionAddStates(StatesGroup):
    choosing_test_type = State()
    waiting_photo = State()
    waiting_answer = State()
    confirming_question = State()


class UserRegistrationStates(StatesGroup):
    waiting_name = State()
    waiting_city = State()


class TestSelectionStates(StatesGroup):
    choosing_test_type = State()


class AdminEditStates(StatesGroup):
    editing_image = State()
    editing_answer = State()


class TestStates(StatesGroup):
    passing_test = State()


# =========================
# KEYBOARDS
# =========================

user_keyboard = ReplyKeyboardMarkup(
    keyboard=[
        [KeyboardButton(text="📝 Пройти тест")],
    ],
    resize_keyboard=True,
)

moderator_keyboard = ReplyKeyboardMarkup(
    keyboard=[
        [KeyboardButton(text="📝 Пройти тест")],
        [KeyboardButton(text="➕ Добавить вопрос")],
        [KeyboardButton(text="📢 Уведомление")],
    ],
    resize_keyboard=True,
)

statistician_keyboard = ReplyKeyboardMarkup(
    keyboard=[
        [KeyboardButton(text="📊 Статистика")],
        [KeyboardButton(text="📁 Экспорт Excel")],
    ],
    resize_keyboard=True,
)

admin_keyboard = ReplyKeyboardMarkup(
    keyboard=[
        [KeyboardButton(text="📝 Пройти тест")],
        [KeyboardButton(text="➕ Добавить вопрос")],
        [KeyboardButton(text="👀 Просмотреть тест")],
        [KeyboardButton(text="📊 Статистика")],
        [KeyboardButton(text="📁 Экспорт Excel")],
        [KeyboardButton(text="📢 Уведомление")],
    ],
    resize_keyboard=True,
)


# =========================
# HELPERS
# =========================

async def create_db():
    logger.info("📦 Creating database tables...")
    try:
        sql_script = INIT_SCHEMA_SQL.read_text(encoding="utf-8")

        async with engine.begin() as conn:
            for statement in sql_script.split(";"):
                statement = statement.strip()
                if statement:
                    await conn.execute(text(statement))
        logger.info("✅ Database tables ready")
    except Exception as e:
        logger.error(f"❌ Database creation error: {e}")
        raise


async def is_admin(user_id: int):
    return user_id in ADMIN_IDS


async def is_moderator(user_id: int):
    return user_id in MODERATOR_IDS


async def is_statistician(user_id: int):
    return user_id in STATISTICIAN_IDS


async def is_admin_or_moderator(user_id: int):
    return user_id in ADMIN_IDS or user_id in MODERATOR_IDS


async def can_view_reports(user_id: int):
    return user_id in ADMIN_IDS or user_id in STATISTICIAN_IDS


async def get_user_by_tg_id(
    session: AsyncSession,
    telegram_id: int,
):
    result = await session.execute(
        select(User).where(
            User.telegram_id == telegram_id
        )
    )

    return result.scalar_one_or_none()


async def get_active_test(session: AsyncSession):

    result = await session.execute(
        select(Test).where(
            Test.test_date == date.today(),
            Test.active == True,
            Test.approved == True,
        )
    )

    return result.scalar_one_or_none()


async def get_test_for_passing(session: AsyncSession):
    """Получить тест для прохождения (без проверки одобрения)"""
    
    result = await session.execute(
        select(Test).where(
            Test.test_date == date.today(),
            Test.active == True,
        )
    )

    return result.scalars().first()


async def get_today_questions(
    session: AsyncSession,
    test_id: int,
    test_type: str = None,
):
    query = select(Question).where(Question.test_id == test_id)
    
    if test_type:
        query = query.where(Question.test_type == test_type)
    
    query = query.order_by(Question.question_order)
    
    result = await session.execute(query)

    return result.scalars().all()


async def already_completed_today(
    session: AsyncSession,
    user_id: int,
    test_id: int,
    test_type: str = "control",
):
    result = await session.execute(
        select(TestSession).where(
            TestSession.user_id == user_id,
            TestSession.test_id == test_id,
            TestSession.test_type == test_type,
            TestSession.completed == True,
        )
    )

    return result.scalar_one_or_none() is not None


async def get_test_session(
    session: AsyncSession,
    user_id: int,
    test_id: int,
    test_type: str,
):
    result = await session.execute(
        select(TestSession).where(
            TestSession.user_id == user_id,
            TestSession.test_id == test_id,
            TestSession.test_type == test_type,
        )
    )

    return result.scalar_one_or_none()


async def get_current_attempt(
    session: AsyncSession,
    test_session: TestSession,
):
    result = await session.execute(
        select(TestAttempt)
        .where(
            TestAttempt.session_id == test_session.id,
            TestAttempt.completed_at.is_(None),
        )
        .order_by(TestAttempt.attempt_number.desc())
    )

    attempt = result.scalars().first()

    if attempt:
        return attempt

    attempt = TestAttempt(
        session_id=test_session.id,
        attempt_number=test_session.attempts or 1,
        passed=False,
    )
    session.add(attempt)
    await session.flush()

    return attempt


async def get_or_create_test_session(
    session: AsyncSession,
    user_id: int,
    test_id: int,
    test_type: str,
):
    test_session = await get_test_session(
        session,
        user_id,
        test_id,
        test_type,
    )

    if test_session:
        if not test_session.completed:
            await get_current_attempt(session, test_session)
        return test_session

    test_session = TestSession(
        user_id=user_id,
        test_id=test_id,
        test_type=test_type,
        current_question=0,
        completed=False,
        attempts=1,
    )
    session.add(test_session)
    await session.flush()

    session.add(
        TestAttempt(
            session_id=test_session.id,
            attempt_number=1,
            passed=False,
        )
    )

    return test_session


async def is_test_approved(
    session: AsyncSession,
    test_id: int,
) -> bool:
    result = await session.execute(
        select(Test.approved).where(
            Test.id == test_id
        )
    )

    return bool(result.scalar_one_or_none())


def format_russian_date(value: date) -> str:
    months = {
        1: "января",
        2: "февраля",
        3: "марта",
        4: "апреля",
        5: "мая",
        6: "июня",
        7: "июля",
        8: "августа",
        9: "сентября",
        10: "октября",
        11: "ноября",
        12: "декабря",
    }

    return f"{value.day} {months[value.month]} {value.year} года"


def get_test_type_name(test_type: str) -> str:
    return "Контроль" if test_type == "control" else "Опыт"


# =========================
# START
# =========================

@dp.message(Command("start"))
async def start_handler(message: Message, state: FSMContext):
    logger.info(f"👤 User {message.from_user.id} started bot")

    async with async_session() as session:

        user = await get_user_by_tg_id(
            session,
            message.from_user.id,
        )

        if not user:

            user = User(
                telegram_id=message.from_user.id,
                full_name="Unknown",
                is_whitelisted=False,
                city="",
            )

            session.add(user)
            await session.commit()

        if (
            not user.is_whitelisted
            and message.from_user.id not in ADMIN_IDS
            and message.from_user.id not in MODERATOR_IDS
            and message.from_user.id not in STATISTICIAN_IDS
        ):
            await state.clear()
            await message.answer(
                "⛔ У вас нет доступа"
            )
            return

        # Если имя неполное или пусто - запрашиваем
        if not user.full_name or user.full_name == "Unknown":
            await state.set_state(UserRegistrationStates.waiting_name)
            msg = await message.answer("👤 Как вас зовут?")
            await state.update_data(instruction_message_id=msg.message_id)
            return

        # Если город не заполнен - запрашиваем
        if not user.city:
            await state.set_state(UserRegistrationStates.waiting_city)
            msg = await message.answer("🏙️ Из какого вы города?")
            await state.update_data(instruction_message_id=msg.message_id)
            return

        if message.from_user.id in ADMIN_IDS:
            keyboard = admin_keyboard
        elif message.from_user.id in MODERATOR_IDS:
            keyboard = moderator_keyboard
        elif message.from_user.id in STATISTICIAN_IDS:
            keyboard = statistician_keyboard
        else:
            keyboard = user_keyboard

        await message.answer(
            "✅ Добро пожаловать",
            reply_markup=keyboard,
        )


# =========================
# USER REGISTRATION
# =========================

@dp.message(UserRegistrationStates.waiting_name)
async def handle_user_name(message: Message, state: FSMContext):
    if not message.text:
        return

    await state.update_data(user_name=message.text)
    await state.set_state(UserRegistrationStates.waiting_city)

    data = await state.get_data()
    if data.get("instruction_message_id"):
        try:
            await message.bot.delete_message(
                message.chat.id,
                data["instruction_message_id"]
            )
        except:
            pass

    await message.delete()

    msg = await message.answer("🏙️ Из какого вы города?")
    await state.update_data(instruction_message_id=msg.message_id)


@dp.message(UserRegistrationStates.waiting_city)
async def handle_user_city(message: Message, state: FSMContext):
    if not message.text:
        return

    data = await state.get_data()
    user_name = data.get("user_name", "")
    city = message.text

    async with async_session() as session:
        user = await get_user_by_tg_id(
            session,
            message.from_user.id,
        )

        if user_name:
            user.full_name = user_name
        user.city = city

        await session.commit()

    await state.clear()
    await message.delete()

    if data.get("instruction_message_id"):
        try:
            await message.bot.delete_message(
                message.chat.id,
                data["instruction_message_id"]
            )
        except:
            pass

    if message.from_user.id in ADMIN_IDS:
        keyboard = admin_keyboard
    elif message.from_user.id in MODERATOR_IDS:
        keyboard = moderator_keyboard
    elif message.from_user.id in STATISTICIAN_IDS:
        keyboard = statistician_keyboard
    else:
        keyboard = user_keyboard

    await message.answer(
        "✅ Добро пожаловать",
        reply_markup=keyboard,
    )


# =========================
# BUTTONS
# =========================

@dp.message(F.text == "📝 Пройти тест")
async def start_test_button(
    message: Message,
    state: FSMContext,
):
    await start_test(message, state)


@dp.message(F.text == "➕ Добавить вопрос")
async def add_question_button(
    message: Message,
    state: FSMContext,
):
    await add_question_start(message, state)


@dp.message(F.text == "📊 Статистика")
async def results_button(message: Message, state: FSMContext):
    if not await can_view_reports(message.from_user.id):
        await message.answer("⛔ У вас нет доступа")
        return

    await state.clear()
    await show_statistics_menu(message)


@dp.message(F.text == "📁 Экспорт Excel")
async def export_button(message: Message):
    await export_handler(message)


@dp.message(F.text == "📢 Уведомление")
async def notification_button(message: Message):
    await notification_handler(message)


# =========================
# TEST TYPE SELECTION
# =========================

async def choose_test_type(message: Message, state: FSMContext):
    await state.set_state(TestSelectionStates.choosing_test_type)

    kb = InlineKeyboardBuilder()
    kb.button(text="📋 Контроль", callback_data="test_type:control")
    kb.button(text="📚 Опыт", callback_data="test_type:experience")
    kb.adjust(1)

    msg = await message.answer(
        "Выберите тип теста:",
        reply_markup=kb.as_markup()
    )
    await state.update_data(test_selection_message_id=msg.message_id)


@dp.callback_query(F.data.startswith("test_type:"))
async def handle_test_type(callback: CallbackQuery, state: FSMContext):
    test_type = callback.data.split(":")[1]

    await callback.answer()
    await callback.message.delete()

    await state.update_data(test_type=test_type)
    await state.set_state(TestStates.passing_test)

    # Используем callback.from_user.id для получения пользователя
    # и отправляем новое сообщение
    await start_test_from_callback(callback, state, test_type)


async def start_test_from_callback(
    callback: CallbackQuery,
    state: FSMContext,
    test_type: str = "control",
):
    async with async_session() as session:

        user = await get_user_by_tg_id(
            session,
            callback.from_user.id,
        )

        if (
            not user
            or (
                not user.is_whitelisted
                and callback.from_user.id not in ADMIN_IDS
            )
        ):
            await callback.message.answer(
                "⛔ Нет доступа"
            )
            return

        test = await get_active_test(session)

        if not test:
            await callback.message.answer(
               "❌ На сегодня тест отсутствует или не одобрен"
            )
            return

        completed = await already_completed_today(
            session,
            user.id,
            test.id,
            test_type,
        )

        if completed:
            await callback.message.answer(
                f"❌ Вы уже проходили этот тест ({test_type}) сегодня"
            )
            return

        questions = await get_today_questions(
            session,
            test.id,
            test_type,
        )

        if not questions:
            await callback.message.answer(
                "❌ Вопросы отсутствуют"
            )
            return

        new_session = await get_or_create_test_session(
            session,
            user_id=user.id,
            test_id=test.id,
            test_type=test_type,
        )

        await session.commit()

        # Инициализируем прогресс для этого типа теста
        # Сохраняем test_type в состояние для использования в send_question
        await state.update_data(test_type=test_type)

        await send_question(
            callback.from_user.id,
            state,
        )


# =========================
# ALLOW
# =========================

@dp.message(Command("allow"))
async def allow_user(message: Message):

    if not await is_admin(message.from_user.id):
        await message.answer("⛔ У вас нет доступа")
        return

    args = message.text.split()

    if len(args) != 2:
        await message.answer(
            "Использование:\n/allow TELEGRAM_ID"
        )
        return

    telegram_id = int(args[1])

    async with async_session() as session:

        user = await get_user_by_tg_id(
            session,
            telegram_id,
        )

        if not user:
            await message.answer(
                "Пользователь ещё не запускал бота"
            )
            return

        user.is_whitelisted = True

        await session.commit()

        await message.answer(
            "✅ Пользователь добавлен"
        )


# =========================
# ADD QUESTION
# =========================

@dp.message(Command("add_question"))
@dp.message(F.text == "➕ Добавить вопрос")
async def add_question_start(
    message: Message,
    state: FSMContext,
):

    if not await is_admin_or_moderator(message.from_user.id):
        await message.answer("⛔ У вас нет доступа")
        return

    async with async_session() as session:

        test = await get_test_for_passing(session)

        if not test:

            test = Test(
                title=f"Тест {date.today()}",
                test_date=date.today(),
                active=True,
            )

            session.add(test)
            await session.commit()

        if test.approved:
            await state.clear()
            await message.answer(
                "🔐 Тест уже опубликован, добавление вопросов закрыто"
            )
            return

        await state.update_data(
            test_id=test.id
        )

    # Запрашиваем тип теста
    await state.set_state(QuestionAddStates.choosing_test_type)

    kb = InlineKeyboardBuilder()
    kb.button(text="📋 Контроль", callback_data="add_test_type:control")
    kb.button(text="📚 Опыт", callback_data="add_test_type:experience")
    kb.adjust(1)

    msg = await message.answer(
        "Выберите тип теста для добавления вопроса:",
        reply_markup=kb.as_markup()
    )

    await state.update_data(instruction_message_id=msg.message_id)


@dp.callback_query(F.data.startswith("add_test_type:"))
async def handle_add_test_type(callback: CallbackQuery, state: FSMContext):
    test_type = callback.data.split(":")[1]

    await callback.answer()
    await callback.message.delete()

    await state.update_data(question_test_type=test_type)
    await state.set_state(QuestionAddStates.waiting_photo)

    msg = await callback.message.answer(
        "📷 Отправьте изображение"
    )

    await state.update_data(
        instruction_message_id=msg.message_id
    )


@dp.callback_query(F.data.startswith("confirm_question:"))
async def handle_confirm_question(callback: CallbackQuery, state: FSMContext):
    action = callback.data.split(":")[1]
    
    await callback.answer()
    await callback.message.delete()
    
    data = await state.get_data()
    
    if action == "no":
        # Переснять - возвращаемся к выбору фото
        await state.set_state(QuestionAddStates.waiting_photo)
        msg = await callback.message.answer(
            "📷 Отправьте новое изображение"
        )
        await state.update_data(
            instruction_message_id=msg.message_id
        )
        return
    
    # action == "yes" - подтверждаем и сохраняем вопрос
    async with async_session() as session:
        if await is_test_approved(session, data["test_id"]):
            await state.clear()
            await callback.message.answer(
                "🔐 Тест уже опубликован, добавление вопросов закрыто"
            )
            return

        result = await session.execute(
            select(Question).where(
                Question.test_id == data["test_id"],
                Question.test_type == data.get("question_test_type", "control"),
            )
        )
        
        questions = result.scalars().all()
        next_order = len(questions) + 1
        
        question = Question(
            test_id=data["test_id"],
            test_type=data.get("question_test_type", "control"),
            question_order=next_order,
            image_file_id=data["photo_id"],
            correct_answer=data["correct_answer"],
        )
        
        session.add(question)
        await session.commit()
    
    kb = InlineKeyboardBuilder()
    kb.button(text="➕ Добавить еще вопрос", callback_data="add_question_more")
    kb.button(text="✅ Закрыть", callback_data="add_question_finish")
    kb.adjust(1)

    await state.update_data(
        test_id=data["test_id"],
        question_test_type=data.get("question_test_type", "control"),
    )

    await callback.message.answer(
        "✅ Вопрос сохранён",
        reply_markup=kb.as_markup(),
    )


@dp.callback_query(F.data == "add_question_more")
async def add_question_more(callback: CallbackQuery, state: FSMContext):
    await callback.answer()
    await callback.message.delete()

    data = await state.get_data()

    async with async_session() as session:
        if await is_test_approved(session, data["test_id"]):
            await state.clear()
            await callback.message.answer(
                "🔐 Тест уже опубликован, добавление вопросов закрыто"
            )
            return

    await state.set_state(QuestionAddStates.waiting_photo)

    msg = await callback.message.answer(
        "📷 Отправьте изображение следующего вопроса"
    )

    await state.update_data(
        instruction_message_id=msg.message_id
    )


@dp.callback_query(F.data == "add_question_finish")
async def add_question_finish(callback: CallbackQuery, state: FSMContext):
    await callback.answer()
    await callback.message.delete()

    await state.clear()

    await callback.message.answer(
        "✅ Добавление вопросов завершено"
    )


# =========================
# PHOTO HANDLER
# =========================

@dp.message(F.photo)
async def photo_handler(
    message: Message,
    state: FSMContext,
):

    current_state = await state.get_state()

    # ДОБАВЛЕНИЕ ВОПРОСА
    if current_state == QuestionAddStates.waiting_photo.state:

        photo = message.photo[-1]
        data = await state.get_data()

        async with async_session() as session:
            if await is_test_approved(session, data["test_id"]):
                await state.clear()
                await message.answer(
                    "🔐 Тест уже опубликован, добавление вопросов закрыто"
                )
                return

        await state.update_data(
            photo_id=photo.file_id
        )

        await state.set_state(QuestionAddStates.waiting_answer)

        await message.delete()

        # Удаляем предыдущее инструкционное сообщение
        if data.get("instruction_message_id"):
            try:
                await message.bot.delete_message(
                    message.chat.id,
                    data["instruction_message_id"]
                )
            except:
                pass

        msg = await message.answer(
            "Введите правильный ответ числом"
        )

        await state.update_data(
            instruction_message_id=msg.message_id
        )

        return

    # РЕДАКТИРОВАНИЕ КАРТИНКИ
    if current_state == AdminEditStates.editing_image.state:

        photo = message.photo[-1]

        data = await state.get_data()

        async with async_session() as session:

            result = await session.execute(
                select(Question, Test)
                .join(Test, Test.id == Question.test_id)
                .where(Question.id == data["edit_question_id"])
            )

            row = result.first()

            if not row:
                await state.clear()
                await message.answer("❌ Вопрос не найден")
                return

            q, test = row

            if test.approved:
                await state.clear()
                await message.answer(
                    "🔐 Тест уже опубликован, редактирование закрыто"
                )
                return

            q.image_file_id = photo.file_id

            await session.commit()

        await state.clear()

        await message.delete()

        # Удаляем инструкционное сообщение
        if data.get("instruction_message_id"):
            try:
                await message.bot.delete_message(
                    message.chat.id,
                    data["instruction_message_id"]
                )
            except:
                pass

        await message.answer(
            "✅ Картинка обновлена"
        )

        return





# =========================
# TEST
# =========================

@dp.message(Command("test"))
async def start_test_command(
    message: Message,
    state: FSMContext,
):
    await start_test(message, state)


async def start_test(
    message: Message,
    state: FSMContext,
    test_type: str | None = None,
):
    requested_test_type = test_type
    logger.info(f"📝 User {message.from_user.id} starting test")

    async with async_session() as session:

        user = await get_user_by_tg_id(
            session,
            message.from_user.id,
        )

        if (
            not user
            or (
                not user.is_whitelisted
                and message.from_user.id not in ADMIN_IDS
            )
        ):
            await message.answer(
                "⛔ Нет доступа"
            )
            return

        test = await get_active_test(session)

        if not test:
            await message.answer(
                "❌ На сегодня тест отсутствует или не одобрен"
            )
            return

        if requested_test_type is None:
            for candidate_type in ("control", "experience"):
                completed = await already_completed_today(
                    session,
                    user.id,
                    test.id,
                    candidate_type,
                )
                questions = await get_today_questions(
                    session,
                    test.id,
                    candidate_type,
                )

                if questions and not completed:
                    test_type = candidate_type
                    break

            if test_type is None:
                await message.answer(
                    "✅ Вы уже прошли все доступные тесты сегодня"
                )
                return

        logger.info(f"📝 User {message.from_user.id} starting {test_type} test")

        completed = await already_completed_today(
            session,
            user.id,
            test.id,
            test_type,
        )

        if completed:
            await message.answer(
                f"❌ Вы уже проходили этот тест ({test_type}) сегодня"
            )
            return

        questions = await get_today_questions(
            session,
            test.id,
            test_type,
        )

        if not questions:
            await message.answer(
                "❌ Вопросы отсутствуют"
            )
            return

        new_session = await get_or_create_test_session(
            session,
            user_id=user.id,
            test_id=test.id,
            test_type=test_type,
        )

        await session.commit()

        # Инициализируем прогресс для этого типа теста
        # Сохраняем test_type в состояние для использования в send_question
        await state.update_data(test_type=test_type)

        await state.set_state(
            TestStates.passing_test
        )

        if test_type == "control":
            await message.answer("📋 Начинаем контроль")
        else:
            await message.answer("📚 Начинаем опыт")

        await send_question(
            message.from_user.id,
            state,
        )


# =========================
# SEND QUESTION
# =========================

async def send_question(
    user_tg_id: int,
    state: FSMContext,
):

    async with async_session() as session:

        test = await get_test_for_passing(session)

        if not test:
            return

        data = await state.get_data()
        test_type = data.get("test_type", "control")

        user = await get_user_by_tg_id(
            session,
            user_tg_id,
        )

        if not user:
            return

        test_session = await get_test_session(
            session,
            user.id,
            test.id,
            test_type,
        )

        if not test_session or test_session.completed:
            return

        attempt = await get_current_attempt(session, test_session)

        questions = await get_today_questions(
            session,
            test.id,
            test_type,
        )

        index = test_session.current_question or 0

        if index >= len(questions):
            answers_result = await session.execute(
                select(Answer, Question)
                .join(Question, Question.id == Answer.question_id)
                .where(Answer.attempt_id == attempt.id)
                .where(Question.test_id == test.id)
                .where(Question.test_type == test_type)
            )
            answers_data = answers_result.all()

            all_correct = (
                len(answers_data) == len(questions)
                and all(answer.is_correct for answer, _ in answers_data)
            )

            now = datetime.now(timezone.utc)

            if not all_correct:
                failed_attempt_number = attempt.attempt_number
                wrong_answers = [
                    (answer, question)
                    for answer, question in answers_data
                    if not answer.is_correct
                ]

                attempt.completed_at = now
                attempt.passed = False
                test_session.attempts = (test_session.attempts or 1) + 1
                test_session.current_question = 0

                session.add(
                    TestAttempt(
                        session_id=test_session.id,
                        attempt_number=test_session.attempts,
                        passed=False,
                    )
                )

                await session.commit()

                data = await state.get_data()
                if data.get("last_question_message_id"):
                    try:
                        await bot.delete_message(
                            user_tg_id,
                            data["last_question_message_id"]
                        )
                    except:
                        pass

                logger.warning(f"⚠️ User {user_tg_id} failed test {test.id} ({test_type}), attempt {test_session.attempts}")

                await bot.send_message(
                    user_tg_id,
                    f"❌ Тест не пройден\n\n"
                    f"{get_test_type_name(test_type)}\n"
                    f"Провальная попытка #{failed_attempt_number}\n\n"
                    f"Начнём заново! 🔄"
                )

                if failed_attempt_number == 5 and wrong_answers:
                    await bot.send_message(
                        user_tg_id,
                        "Ниже вопросы, где были ошибки. Правильные ответы не показываю."
                    )

                    for wrong_answer, wrong_question in wrong_answers:
                        await bot.send_photo(
                            user_tg_id,
                            wrong_question.image_file_id,
                            caption=(
                                f"❌ {get_test_type_name(test_type)}\n"
                                f"Вопрос {wrong_question.question_order}\n"
                                f"Ваш ответ: {wrong_answer.answer}"
                            ),
                        )

                await state.update_data(test_type=test_type)

                await send_question(
                    user_tg_id,
                    state,
                )
                return

            attempt.completed_at = now
            attempt.passed = True
            test_session.completed = True
            test_session.completed_at = now
            test_session.current_question = len(questions)

            await session.commit()

            data = await state.get_data()
            if data.get("last_question_message_id"):
                try:
                    await bot.delete_message(
                        user_tg_id,
                        data["last_question_message_id"]
                    )
                except:
                    pass

            logger.info(f"✅ User {user_tg_id} completed test {test.id} ({test_type}) in {test_session.attempts} attempt(s)")

            if test_type == "control":
                experience_questions = await get_today_questions(
                    session,
                    test.id,
                    "experience",
                )
                experience_completed = await already_completed_today(
                    session,
                    user.id,
                    test.id,
                    "experience",
                )

                if experience_questions and not experience_completed:
                    await get_or_create_test_session(
                        session,
                        user.id,
                        test.id,
                        "experience",
                    )
                    await session.commit()

                    await bot.send_message(
                        user_tg_id,
                        f"✅ Контроль завершён успешно\n\n"
                        f"Количество попыток: {test_session.attempts}\n\n"
                        f"📚 Переходим к опыту"
                    )

                    await state.update_data(test_type="experience")
                    await send_question(
                        user_tg_id,
                        state,
                    )
                    return

            await bot.send_message(
                user_tg_id,
                f"✅ {get_test_type_name(test_type)} завершён успешно!\n\n"
                f"Количество попыток: {test_session.attempts}"
            )

            await state.clear()
            return

        question = questions[index]

        data = await state.get_data()
        if data.get("last_question_message_id"):
            try:
                await bot.delete_message(
                    user_tg_id,
                    data["last_question_message_id"]
                )
            except:
                pass

        sent_message = await bot.send_photo(
            user_tg_id,
            question.image_file_id,
            caption=(
                f"{get_test_type_name(test_type)}\n"
                f"Вопрос {index + 1} "
                f"из {len(questions)}\n\n"
                f"Введите ответ сообщением"
            ),
        )

        await state.update_data(
            last_question_message_id=sent_message.message_id,
            current_question_id=question.id,
            test_type=test_type,
        )


# =========================
# TEST VIEW
# =========================

@dp.message(F.text == "👀 Просмотреть тест")
async def view_test_button(message: Message, state: FSMContext):
    
    if not await is_admin(message.from_user.id):
        await message.answer("⛔ У вас нет доступа")
        return

    # Запрашиваем тип теста
    await state.set_state(TestSelectionStates.choosing_test_type)

    kb = InlineKeyboardBuilder()
    kb.button(text="📋 Контроль", callback_data="view_test_type:control")
    kb.button(text="📚 Опыт", callback_data="view_test_type:experience")
    kb.adjust(1)

    msg = await message.answer(
        "Выберите тип теста для просмотра:",
        reply_markup=kb.as_markup()
    )


@dp.callback_query(F.data.startswith("view_test_type:"))
async def handle_view_test_type(callback: CallbackQuery, state: FSMContext):
    test_type = callback.data.split(":")[1]

    await callback.answer()
    await callback.message.delete()

    async with async_session() as session:

        test = await get_test_for_passing(session)

        if not test:
            await callback.message.answer(
                "❌ Тест не найден"
            )
            return

        questions = await get_today_questions(
            session,
            test.id,
            test_type,
        )

        if not questions:
            await callback.message.answer(
                f"❌ Вопросов нет в типе '{test_type}'"
            )
            return

        kb = InlineKeyboardBuilder()

        for q in questions:
            kb.button(
                text=f"Вопрос {q.question_order}",
                callback_data=f"admin_q:{q.id}"
            )

        kb.adjust(3)

        await callback.message.answer(
            f"📋 Вопросы теста ({test_type}):",
            reply_markup=kb.as_markup()
        )


@dp.callback_query(F.data.startswith("preview_test_type:"))
async def handle_preview_test_type(callback: CallbackQuery):
    test_type = callback.data.split(":")[1]

    await callback.answer()

    async with async_session() as session:

        test = await get_test_for_passing(session)

        if not test:
            await callback.message.answer(
                "❌ Тест не найден"
            )
            return

        questions = await get_today_questions(
            session,
            test.id,
            test_type,
        )

        if not questions:
            await callback.message.answer(
                f"❌ Вопросов нет в типе '{test_type}'"
            )
            return

        kb = InlineKeyboardBuilder()

        for q in questions:
            kb.button(
                text=f"Вопрос {q.question_order}",
                callback_data=f"admin_q:{q.id}"
            )

        kb.adjust(3)

        await callback.message.answer(
            f"📋 Проверка теста ({test_type}):",
            reply_markup=kb.as_markup()
        )


# =========================
# ADMIN QUESTION VIEW
# =========================

@dp.callback_query(F.data.startswith("admin_q:"))
async def admin_view_question(
    callback: CallbackQuery
):

    if not await is_admin_or_moderator(callback.from_user.id):
        return

    question_id = int(
        callback.data.split(":")[1]
    )

    async with async_session() as session:

        result = await session.execute(
            select(Question).where(
                Question.id == question_id
            )
        )

        q = result.scalar_one()

        test_result = await session.execute(
            select(Test).where(
                Test.id == q.test_id
            )
        )
        test = test_result.scalar_one()

    kb = InlineKeyboardBuilder()

    if not test.approved:
        kb.button(
            text="✏️ Картинка",
            callback_data=f"edit_img:{q.id}"
        )

        kb.button(
            text="✏️ Ответ",
            callback_data=f"edit_ans:{q.id}"
        )

        kb.adjust(1)

    await callback.answer()

    await callback.message.answer_photo(
        q.image_file_id,
        caption=(
            f"📌 Вопрос {q.question_order}\n"
            f"🔒 Ответ: {q.correct_answer}"
            + ("\n\n🔐 Тест уже опубликован, редактирование закрыто" if test.approved else "")
        ),
        reply_markup=kb.as_markup() if not test.approved else None
    )


# =========================
# PROCESS ANSWER
# =========================

async def save_test_answer(
    user_tg_id: int,
    state: FSMContext,
    question_id: int,
    answer: int,
) -> bool:
    async with async_session() as session:

        user = await get_user_by_tg_id(
            session,
            user_tg_id,
        )

        question_result = await session.execute(
            select(Question).where(
                Question.id == question_id
            )
        )
        question = question_result.scalar_one_or_none()

        if not user or not question:
            return False

        test_session = await get_test_session(
            session,
            user.id,
            question.test_id,
            question.test_type,
        )

        if not test_session or test_session.completed:
            return False

        attempt = await get_current_attempt(session, test_session)

        existing_result = await session.execute(
            select(Answer).where(
                Answer.attempt_id == attempt.id,
                Answer.question_id == question.id,
            )
        )
        existing_answer = existing_result.scalar_one_or_none()

        is_correct = answer == question.correct_answer

        if existing_answer:
            existing_answer.answer = answer
            existing_answer.is_correct = is_correct
            existing_answer.correct_answer_at_time = question.correct_answer
        else:
            session.add(
                Answer(
                    attempt_id=attempt.id,
                    user_id=user.id,
                    question_id=question.id,
                    answer=answer,
                    is_correct=is_correct,
                    correct_answer_at_time=question.correct_answer,
                )
            )
            test_session.current_question = (test_session.current_question or 0) + 1

        await session.commit()
        await state.update_data(test_type=question.test_type)

    return True


@dp.callback_query(F.data.startswith("answer:"))
async def process_answer(
    callback: CallbackQuery,
    state: FSMContext,
):

    parts = callback.data.split(":")

    question_id = int(parts[1])

    answer = int(parts[2])

    saved = await save_test_answer(
        callback.from_user.id,
        state,
        question_id,
        answer,
    )

    if not saved:
        await callback.answer()
        return

    await callback.message.delete()

    await callback.answer()

    await send_question(
        callback.from_user.id,
        state,
    )


@dp.callback_query(F.data.startswith("confirm_test_answer:"))
async def confirm_test_answer(callback: CallbackQuery, state: FSMContext):
    action = callback.data.split(":")[1]

    await callback.answer()
    await callback.message.delete()

    data = await state.get_data()

    if action == "no":
        await state.update_data(
            pending_question_id=None,
            pending_answer=None,
        )
        await callback.message.answer(
            "Введите ответ еще раз"
        )
        return

    question_id = data.get("pending_question_id")
    answer = data.get("pending_answer")

    if question_id is None or answer is None:
        await callback.message.answer(
            "❌ Ответ не найден. Введите ответ еще раз"
        )
        return

    saved = await save_test_answer(
        callback.from_user.id,
        state,
        int(question_id),
        int(answer),
    )

    await state.update_data(
        pending_question_id=None,
        pending_answer=None,
    )

    if not saved:
        await callback.message.answer(
            "❌ Не удалось сохранить ответ. Попробуйте еще раз"
        )
        return

    await send_question(
        callback.from_user.id,
        state,
    )


# =========================
# IMAGE EDIT
# =========================

@dp.callback_query(F.data.startswith("edit_img:"))
async def edit_image_start(
    callback: CallbackQuery,
    state: FSMContext,
):

    if not await is_admin(callback.from_user.id):
        await callback.answer(
            "⛔ У вас нет доступа к редактированию. Напишите администратору @dezztape",
            show_alert=True,
        )
        return

    question_id = int(
        callback.data.split(":")[1]
    )

    async with async_session() as session:
        result = await session.execute(
            select(Question, Test)
            .join(Test, Test.id == Question.test_id)
            .where(Question.id == question_id)
        )
        row = result.first()

        if not row:
            await callback.answer(
                "Вопрос не найден",
                show_alert=True,
            )
            return

        _, test = row

        if test.approved:
            await callback.answer(
                "Тест уже опубликован, редактирование закрыто",
                show_alert=True,
            )
            return

    await state.update_data(
        edit_question_id=question_id
    )

    await state.set_state(
        AdminEditStates.editing_image
    )

    await callback.answer()

    msg = await callback.message.answer(
        "📸 Отправьте новую картинку"
    )

    await state.update_data(
        instruction_message_id=msg.message_id
    )


# =========================
# ANSWER EDIT
# =========================

@dp.callback_query(F.data.startswith("edit_ans:"))
async def edit_answer_start(
    callback: CallbackQuery,
    state: FSMContext,
):

    if not await is_admin(callback.from_user.id):
        await callback.answer(
            "⛔ У вас нет доступа к редактированию. Напишите администратору @dezztape",
            show_alert=True,
        )
        return

    question_id = int(
        callback.data.split(":")[1]
    )

    async with async_session() as session:
        result = await session.execute(
            select(Question, Test)
            .join(Test, Test.id == Question.test_id)
            .where(Question.id == question_id)
        )
        row = result.first()

        if not row:
            await callback.answer(
                "Вопрос не найден",
                show_alert=True,
            )
            return

        _, test = row

        if test.approved:
            await callback.answer(
                "Тест уже опубликован, редактирование закрыто",
                show_alert=True,
            )
            return

    await state.update_data(
        edit_question_id=question_id
    )

    await state.set_state(
        AdminEditStates.editing_answer
    )

    await callback.answer()

    msg = await callback.message.answer(
        "✏️ Введите новый правильный ответ числом"
    )

    await state.update_data(
        instruction_message_id=msg.message_id
    )


# =========================
# RESULTS
# =========================

@dp.callback_query(F.data.startswith("stats_type:"))
async def handle_stats_type(callback: CallbackQuery):
    if not await can_view_reports(callback.from_user.id):
        await callback.answer(
            "У вас нет доступа",
            show_alert=True,
        )
        return

    test_type = callback.data.split(":")[1]
    
    await callback.answer()
    await callback.message.delete()
    
    await show_statistics(callback.message, test_type)


@dp.message(Command("results"))
async def results_command(message: Message, state: FSMContext):
    if not await can_view_reports(message.from_user.id):
        await message.answer("⛔ У вас нет доступа")
        return

    await state.clear()
    await show_statistics_menu(message)


async def show_statistics_menu(message: Message):
    kb = InlineKeyboardBuilder()
    kb.button(text="📅 Сегодня", callback_data="stats_date:today")
    kb.button(text="↩️ Вчера", callback_data="stats_date:yesterday")
    kb.button(text="🗓 Доступные дни", callback_data="stats_available_days")
    kb.adjust(1)

    await message.answer(
        "Выберите день для просмотра статистики:",
        reply_markup=kb.as_markup(),
    )


@dp.callback_query(F.data.startswith("stats_date:"))
async def handle_stats_date(callback: CallbackQuery):
    if not await can_view_reports(callback.from_user.id):
        await callback.answer(
            "У вас нет доступа",
            show_alert=True,
        )
        return

    period = callback.data.split(":")[1]
    stats_date = date.today()

    if period == "yesterday":
        stats_date = date.today() - timedelta(days=1)

    await callback.answer()
    await callback.message.delete()

    async with async_session() as session:
        result = await session.execute(
            select(Test).where(
                Test.test_date == stats_date,
            )
        )
        test = result.scalar_one_or_none()

    if not test:
        await callback.message.answer(
            f"❌ Тест за {format_russian_date(stats_date)} не найден"
        )
        return

    await show_statistics(callback.message, test_id=test.id)


@dp.callback_query(F.data == "stats_available_days")
async def handle_stats_available_days(callback: CallbackQuery):
    if not await can_view_reports(callback.from_user.id):
        await callback.answer(
            "У вас нет доступа",
            show_alert=True,
        )
        return

    await callback.answer()
    await callback.message.delete()

    async with async_session() as session:
        result = await session.execute(
            select(Test)
            .order_by(Test.test_date.desc())
            .limit(30)
        )
        tests = result.scalars().all()

    if not tests:
        await callback.message.answer(
            "❌ Доступных дней для статистики нет"
        )
        return

    kb = InlineKeyboardBuilder()
    for test in tests:
        kb.button(
            text=format_russian_date(test.test_date),
            callback_data=f"stats_test:{test.id}",
        )
    kb.adjust(1)

    await callback.message.answer(
        "Выберите день для просмотра статистики:",
        reply_markup=kb.as_markup(),
    )


@dp.callback_query(F.data.startswith("stats_test:"))
async def handle_stats_test(callback: CallbackQuery):
    if not await can_view_reports(callback.from_user.id):
        await callback.answer(
            "У вас нет доступа",
            show_alert=True,
        )
        return

    test_id = int(callback.data.split(":")[1])

    await callback.answer()
    await callback.message.delete()

    await show_statistics(callback.message, test_id=test_id)


async def show_statistics(
    message: Message,
    test_type: str | None = None,
    test_id: int | None = None,
):

    async with async_session() as session:

        if test_id is None:
            test = await get_test_for_passing(session)
        else:
            result = await session.execute(
                select(Test).where(
                    Test.id == test_id,
                )
            )
            test = result.scalar_one_or_none()

        if not test:
            await message.answer(
                "❌ Тест не найден"
            )
            return

        excluded_stats_ids = set(ADMIN_IDS + STATISTICIAN_IDS)

        # Получаем сотрудников без админов и роли просмотра статистики
        users_result = await session.execute(
            select(User).order_by(User.full_name)
        )
        all_users = [
            user
            for user in users_result.scalars().all()
            if user.telegram_id not in excluded_stats_ids
        ]

        if not all_users:
            await message.answer(
                "❌ Сотрудников не найдено"
            )
            return

        test_types = (
            [(test_type, "Контроль" if test_type == "control" else "Опыт")]
            if test_type
            else [
                ("control", "Контроль"),
                ("experience", "Опыт"),
            ]
        )

        user_stats = {
            user.id: {
                "user": user,
                "types": {
                    current_type: {
                        "total_questions": 0,
                        "correct_answers": 0,
                        "attempts": 1,
                        "status": "❌ Не начат",
                    }
                    for current_type, _ in test_types
                },
            }
            for user in all_users
        }

        for current_type, _ in test_types:
            sessions_result = await session.execute(
                select(TestSession).where(
                    TestSession.test_id == test.id,
                    TestSession.test_type == current_type,
                )
            )
            test_sessions = sessions_result.scalars().all()

            for ts in test_sessions:
                if ts.user_id not in user_stats:
                    continue

                stats = user_stats[ts.user_id]["types"][current_type]
                stats["status"] = "✅ Завершен" if ts.completed else "⏳ В процессе"
                stats["attempts"] = ts.attempts

                attempt_query = select(TestAttempt).where(
                    TestAttempt.session_id == ts.id,
                )

                if ts.completed:
                    attempt_query = attempt_query.where(
                        TestAttempt.passed == True
                    )
                else:
                    attempt_query = attempt_query.where(
                        TestAttempt.completed_at.is_(None)
                    )

                attempt_result = await session.execute(
                    attempt_query.order_by(TestAttempt.attempt_number.desc())
                )
                attempt = attempt_result.scalars().first()

                if not attempt:
                    continue

                answers_result = await session.execute(
                    select(Answer, Question)
                    .join(Question, Question.id == Answer.question_id)
                    .where(Answer.attempt_id == attempt.id)
                    .where(Question.test_id == test.id)
                    .where(Question.test_type == current_type)
                )
                answers_data = answers_result.all()

                for answer, question in answers_data:
                    stats["total_questions"] += 1
                    if answer.is_correct:
                        stats["correct_answers"] += 1

        text = "📊 Статистика\n" if test_type is None else f"📊 Статистика - {test_types[0][1]}\n"
        text += f"Тест №{test.id} от {format_russian_date(test.test_date)}\n"
        text += "=" * 40 + "\n\n"

        for user_id in sorted(user_stats.keys()):
            user_data = user_stats[user_id]
            user = user_data["user"]

            text += f"👤 {user.full_name}\n"
            text += f"📍 {user.city or '-'}\n"

            for current_type, type_name in test_types:
                stats = user_data["types"][current_type]
                total = stats["total_questions"]
                correct = stats["correct_answers"]
                percent = (correct / total * 100) if total > 0 else 0
                type_icon = "📋" if current_type == "control" else "🧪"

                text += f"{type_icon} <b>{type_name}</b>\n"

                if total > 0:
                    text += f"✅ Правильных: {correct}/{total} ({percent:.0f}%)\n"
                    text += f"🔄 Попыток: {stats['attempts']}\n"

                text += f"📌 {stats['status']}\n"

            text += "-" * 40 + "\n\n"

        if len(text) > 4000:
            text = text[:4000]

        await message.answer(text)




# =========================

async def notification_handler(message: Message):

    if not await is_admin_or_moderator(message.from_user.id):
        await message.answer(
            "⛔ У вас нет доступа"
        )
        return

    async with async_session() as session:

        # Получаем сегодняшний тест без проверки approved
        result = await session.execute(
            select(Test).where(
                Test.test_date == date.today(),
                Test.active == True,
            )
        )
        test = result.scalars().first()

        if not test:
            await message.answer(
                "❌ Тест не найден"
            )
            return

        # Проверяем, одобрен ли тест
        if test.approved:
            await message.answer(
                "ℹ️ Тест уже одобрен и опубликован"
            )
            return

        kb = InlineKeyboardBuilder()
        kb.button(text="👀 Проверить контроль", callback_data="preview_test_type:control")
        kb.button(text="👀 Проверить опыт", callback_data="preview_test_type:experience")
        kb.button(text="📢 Отправить уведомление", callback_data="publish_test_notification")
        kb.button(text="❌ Отмена", callback_data="cancel_test_notification")
        kb.adjust(1)

        await message.answer(
            "Перед отправкой уведомления проверьте тест.\n\n"
            "После отправки уведомления редактирование вопросов будет закрыто.",
            reply_markup=kb.as_markup(),
        )


@dp.callback_query(F.data == "cancel_test_notification")
async def cancel_test_notification(callback: CallbackQuery):
    await callback.answer()
    await callback.message.delete()


@dp.callback_query(F.data == "publish_test_notification")
async def publish_test_notification(callback: CallbackQuery):
    if not await is_admin_or_moderator(callback.from_user.id):
        await callback.answer(
            "У вас нет доступа",
            show_alert=True,
        )
        return

    await callback.answer()
    await callback.message.delete()

    async with async_session() as session:
        result = await session.execute(
            select(Test).where(
                Test.test_date == date.today(),
                Test.active == True,
            )
        )
        test = result.scalars().first()

        if not test:
            await callback.message.answer(
                "❌ Тест не найден"
            )
            return

        if test.approved:
            await callback.message.answer(
                "ℹ️ Тест уже одобрен и опубликован"
            )
            return

        test.approved = True
        test.approved_at = datetime.now(timezone.utc)
        await session.commit()

        service_role_ids = set(
            ADMIN_IDS + MODERATOR_IDS + STATISTICIAN_IDS
        )

        users_result = await session.execute(
            select(User)
        )
        users = [
            user
            for user in users_result.scalars().all()
            if user.telegram_id not in service_role_ids
        ]

        success_count = 0
        for user in users:
            try:
                await bot.send_message(
                    user.telegram_id,
                    f"🔔 Тесты открыты для прохождения!\n\n"
                    f"📋 Контроль качества №{test.id}\n"
                    f"Нажмите '📝 Пройти тест' чтобы начать"
                )
                success_count += 1
            except:
                pass

        service_success_count = 0
        for telegram_id in service_role_ids:
            try:
                await bot.send_message(
                    telegram_id,
                    "🔔 Тесты открыты для прохождения!"
                )
                service_success_count += 1
            except:
                pass

        await callback.message.answer(
            f"✅ Тест одобрен!\n\n"
            f"📨 Уведомления отправлены: {success_count} сотрудникам\n"
            f"🔔 Служебные уведомления отправлены: {service_success_count}"
        )


# =========================

@dp.message(Command("export"))
@dp.message(F.text == "📁 Экспорт Excel")
async def export_handler(message: Message):

    if not await can_view_reports(message.from_user.id):
        await message.answer(
            "⛔ У вас нет доступа"
        )
        return

    kb = InlineKeyboardBuilder()
    kb.button(text="📅 Сегодня", callback_data="export_date:today")
    kb.button(text="↩️ Вчера", callback_data="export_date:yesterday")
    kb.button(text="🗓 Доступные дни", callback_data="export_available_days")
    kb.adjust(1)

    await message.answer(
        "Выберите день для выгрузки Excel:",
        reply_markup=kb.as_markup(),
    )


@dp.callback_query(F.data.startswith("export_date:"))
async def handle_export_date(callback: CallbackQuery):
    if not await can_view_reports(callback.from_user.id):
        await callback.answer(
            "У вас нет доступа",
            show_alert=True,
        )
        return

    period = callback.data.split(":")[1]
    export_date = date.today()

    if period == "yesterday":
        export_date = date.today() - timedelta(days=1)

    await callback.answer()
    await callback.message.delete()

    async with async_session() as session:
        result = await session.execute(
            select(Test).where(
                Test.test_date == export_date,
            )
        )
        test = result.scalar_one_or_none()

    if not test:
        await callback.message.answer(
            f"❌ Тест за {format_russian_date(export_date)} не найден"
        )
        return

    await export_test_results(callback.message, test.id)


@dp.callback_query(F.data == "export_available_days")
async def handle_export_available_days(callback: CallbackQuery):
    if not await can_view_reports(callback.from_user.id):
        await callback.answer(
            "У вас нет доступа",
            show_alert=True,
        )
        return

    await callback.answer()
    await callback.message.delete()

    async with async_session() as session:
        result = await session.execute(
            select(Test)
            .order_by(Test.test_date.desc())
            .limit(30)
        )
        tests = result.scalars().all()

    if not tests:
        await callback.message.answer(
            "❌ Доступных дней для экспорта нет"
        )
        return

    kb = InlineKeyboardBuilder()
    for test in tests:
        kb.button(
            text=format_russian_date(test.test_date),
            callback_data=f"export_test:{test.id}",
        )
    kb.adjust(1)

    await callback.message.answer(
        "Выберите день для выгрузки Excel:",
        reply_markup=kb.as_markup(),
    )


@dp.callback_query(F.data.startswith("export_test:"))
async def handle_export_test(callback: CallbackQuery):
    if not await can_view_reports(callback.from_user.id):
        await callback.answer(
            "У вас нет доступа",
            show_alert=True,
        )
        return

    test_id = int(callback.data.split(":")[1])

    await callback.answer()
    await callback.message.delete()

    await export_test_results(callback.message, test_id)


async def export_test_results(message: Message, test_id: int):
    async with async_session() as session:
        result = await session.execute(
            select(Test).where(
                Test.id == test_id,
            )
        )
        test = result.scalar_one_or_none()

        if not test:
            await message.answer(
                "❌ Тест не найден"
            )
            return

        # Получаем только завершённые тестовые сессии
        sessions_result = await session.execute(
            select(TestSession).where(
                TestSession.test_id == test.id,
                TestSession.completed == True,
            )
        )
        test_sessions = sessions_result.scalars().all()

        if not test_sessions:
            await message.answer(
                "❌ Нет данных"
            )
            return

        # Структурируем данные для листов сотрудников и сводки ошибок
        user_data = {}
        mistakes_summary = {}

        for ts in test_sessions:
            result_user = await session.execute(
                select(User).where(User.id == ts.user_id)
            )
            user = result_user.scalar_one_or_none()
            if not user:
                continue

            if user.id not in user_data:
                user_data[user.id] = {
                    'user': user,
                    'sessions': {},
                    'attempts': {},
                    'attempt_history': {},
                }

            test_type = ts.test_type

            # Пропускаем, если уже обработали этот тип для этого пользователя
            if test_type in user_data[user.id]['sessions']:
                continue

            user_data[user.id]['sessions'][test_type] = {}
            user_data[user.id]['attempts'][test_type] = ts.attempts
            user_data[user.id]['attempt_history'][test_type] = {}

            attempts_result = await session.execute(
                select(TestAttempt)
                .where(TestAttempt.session_id == ts.id)
                .order_by(TestAttempt.attempt_number)
            )
            attempts = attempts_result.scalars().all()

            passed_attempt = None

            for attempt in attempts:
                if attempt.passed:
                    passed_attempt = attempt

                answers_result = await session.execute(
                    select(Answer, Question)
                    .join(Question, Question.id == Answer.question_id)
                    .where(Answer.attempt_id == attempt.id)
                    .where(Question.test_id == test.id)
                    .where(Question.test_type == test_type)
                    .order_by(Question.question_order)
                )
                answers = answers_result.all()

                attempt_answers = {}
                for answer, question in answers:
                    q_order = question.question_order
                    attempt_answers[q_order] = {
                        "answer": answer.answer,
                        "is_correct": answer.is_correct,
                    }

                    summary_key = (question.test_type, q_order)
                    if summary_key not in mistakes_summary:
                        mistakes_summary[summary_key] = {
                            "test_type": question.test_type,
                            "question_order": q_order,
                            "mistakes": 0,
                            "total": 0,
                            "users": set(),
                        }

                    mistakes_summary[summary_key]["total"] += 1
                    if not answer.is_correct:
                        mistakes_summary[summary_key]["mistakes"] += 1
                        mistakes_summary[summary_key]["users"].add(
                            user.full_name
                        )

                user_data[user.id]['attempt_history'][test_type][
                    attempt.attempt_number
                ] = attempt_answers

            if passed_attempt:
                final_answers_result = await session.execute(
                    select(Answer, Question)
                    .join(Question, Question.id == Answer.question_id)
                    .where(Answer.attempt_id == passed_attempt.id)
                    .where(Question.test_id == test.id)
                    .where(Question.test_type == test_type)
                    .order_by(Question.question_order)
                )
                final_answers = final_answers_result.all()

                for answer, question in final_answers:
                    q_order = question.question_order
                    user_data[user.id]['sessions'][test_type][q_order] = answer.answer

        # Создаем Excel файл
        filename = f"results_{test.id}_{test.test_date.strftime('%d_%m_%Y')}.xlsx"
        test_date_label = format_russian_date(test.test_date)
        wb = Workbook()
        # Удаляем пустой лист только если будут созданы новые
        default_sheet = wb.active
        sheets_created = False

        # Стили
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        center_alignment = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )
        final_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        attempts_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        attempt_title_fill = PatternFill(start_color="B4A7D6", end_color="B4A7D6", fill_type="solid")
        correct_fill = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
        incorrect_fill = PatternFill(start_color="D99694", end_color="D99694", fill_type="solid")
        summary_warning_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
        summary_ok_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

        def make_unique_sheet_name(base_name: str) -> str:
            clean_name = "".join(
                ch for ch in base_name if ch not in r'[]:*?/\\'
            ).strip() or "User"
            clean_name = clean_name[:31]
            if clean_name not in wb.sheetnames:
                return clean_name

            counter = 2
            while True:
                suffix = f" {counter}"
                candidate = f"{clean_name[:31 - len(suffix)]}{suffix}"
                if candidate not in wb.sheetnames:
                    return candidate
                counter += 1

        def write_header_row(ws, row_number: int, headers: list[str]):
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row_number, column=col)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_alignment
                cell.border = border

        # Создаем лист для каждого пользователя
        for user_id in sorted(user_data.keys()):
            user = user_data[user_id]['user']
            sessions = user_data[user_id]['sessions']

            # Ограничиваем длину названия листа до 31 символа
            sheet_name = make_unique_sheet_name(
                user.full_name if user.full_name else "User"
            )
            
            # Если это первый лист, используем дефолтный, иначе создаём новый
            if not sheets_created:
                ws = default_sheet
                ws.title = sheet_name
                sheets_created = True
            else:
                ws = wb.create_sheet(title=sheet_name)

            # Заголовок
            ws['A1'] = f"Контроль качества №{test.id} {test_date_label}"
            ws['A2'] = user.city if user.city else ""

            # Объединяем ячейки для заголовка
            ws.merge_cells('A1:D1')
            ws.merge_cells('A2:D2')
            ws['A1'].font = Font(bold=True, size=12)
            ws['A2'].font = Font(bold=True, size=11)

            # Заголовки таблицы
            write_header_row(
                ws,
                4,
                ["№", "Контроль", "Опыт", "Результат"],
            )

            # Данные
            row = 5
            control_answers = sessions.get('control', {})
            experience_answers = sessions.get('experience', {})

            all_questions = sorted(set(
                list(control_answers.keys()) + 
                list(experience_answers.keys())
            ))

            for q_num in all_questions:
                ws.cell(row=row, column=1, value=q_num)
                
                control_answer = control_answers.get(q_num, "")
                exp_answer = experience_answers.get(q_num, "")

                ws.cell(row=row, column=2, value=control_answer if control_answer else "")
                ws.cell(row=row, column=3, value=exp_answer if exp_answer else "")
                
                # Результат - пусто для каждой строки
                ws.cell(row=row, column=4, value="")

                # Форматирование
                for col in range(1, 5):
                    cell = ws.cell(row=row, column=col)
                    cell.alignment = center_alignment
                    cell.border = border

                row += 1

            # Строка "Итог"
            itog_row = row
            last_question_row = row - 1
            ws.cell(row=itog_row, column=1, value="Итог")
            
            # Сумма Контроля
            ws.cell(row=itog_row, column=2, value=f"=SUM(B5:B{last_question_row})")
            # Сумма Опыта
            ws.cell(row=itog_row, column=3, value=f"=SUM(C5:C{last_question_row})")
            
            # Результат в строке Итог = сумма опыта / сумма контроля
            result_formula = f"=IF(SUM(B5:B{last_question_row})=0,0,SUM(C5:C{last_question_row})/SUM(B5:B{last_question_row}))"
            ws.cell(row=itog_row, column=4, value=result_formula)

            # Форматирование строки Итог
            ws.cell(row=itog_row, column=1).font = Font(bold=True)
            for col in range(1, 5):
                cell = ws.cell(row=itog_row, column=col)
                cell.fill = final_fill
                cell.border = border
                cell.alignment = center_alignment

            attempts_row = itog_row + 1
            ws.cell(row=attempts_row, column=1, value="Попыток")
            ws.cell(row=attempts_row, column=2, value=user_data[user_id]['attempts'].get('control', ""))
            ws.cell(row=attempts_row, column=3, value=user_data[user_id]['attempts'].get('experience', ""))
            ws.cell(row=attempts_row, column=4, value="")

            for col in range(1, 5):
                cell = ws.cell(row=attempts_row, column=col)
                cell.fill = attempts_fill
                cell.border = border
                cell.alignment = center_alignment
                cell.font = Font(bold=True)

            attempt_history = user_data[user_id]['attempt_history']
            all_attempt_numbers = sorted(set(
                list(attempt_history.get('control', {}).keys()) +
                list(attempt_history.get('experience', {}).keys())
            ))

            row = attempts_row + 3
            for attempt_number in all_attempt_numbers:
                ws.merge_cells(
                    start_row=row,
                    start_column=1,
                    end_row=row,
                    end_column=3,
                )
                title_cell = ws.cell(row=row, column=1)
                title_cell.value = f"Попытка #{attempt_number}"
                title_cell.fill = attempt_title_fill
                title_cell.alignment = center_alignment
                title_cell.border = border
                for col in range(2, 4):
                    cell = ws.cell(row=row, column=col)
                    cell.fill = attempt_title_fill
                    cell.border = border

                row += 1
                write_header_row(ws, row, ["№", "Контроль", "Опыт"])

                control_attempt = attempt_history.get('control', {}).get(
                    attempt_number,
                    {},
                )
                experience_attempt = attempt_history.get('experience', {}).get(
                    attempt_number,
                    {},
                )
                attempt_questions = sorted(set(
                    list(control_attempt.keys()) +
                    list(experience_attempt.keys())
                ))

                row += 1
                for q_num in attempt_questions:
                    ws.cell(row=row, column=1, value=q_num)

                    control_data = control_attempt.get(q_num)
                    experience_data = experience_attempt.get(q_num)

                    if control_data:
                        cell = ws.cell(
                            row=row,
                            column=2,
                            value=control_data["answer"],
                        )
                        cell.fill = correct_fill if control_data["is_correct"] else incorrect_fill

                    if experience_data:
                        cell = ws.cell(
                            row=row,
                            column=3,
                            value=experience_data["answer"],
                        )
                        cell.fill = correct_fill if experience_data["is_correct"] else incorrect_fill

                    for col in range(1, 4):
                        cell = ws.cell(row=row, column=col)
                        cell.alignment = center_alignment
                        cell.border = border

                    row += 1

                row += 2

            # Ширина колонок
            ws.column_dimensions['A'].width = 12
            ws.column_dimensions['B'].width = 12
            ws.column_dimensions['C'].width = 12
            ws.column_dimensions['D'].width = 12

        summary_sheet_name = make_unique_sheet_name("Ошибки по вопросам")
        ws = wb.create_sheet(title=summary_sheet_name)

        ws.merge_cells('A1:F1')
        ws['A1'] = f"Сводка ошибок по тесту №{test.id} {test_date_label}"
        ws['A1'].font = Font(bold=True, size=12)
        ws['A1'].alignment = Alignment(horizontal="left", vertical="center")

        write_header_row(
            ws,
            3,
            ["№ вопроса", "Тип", "Ошибок", "Всего ответов", "% ошибок", "Кто ошибся"],
        )

        summary_rows = sorted(
            mistakes_summary.values(),
            key=lambda item: (
                -item["mistakes"],
                item["test_type"],
                item["question_order"],
            )
        )

        row = 4
        for item in summary_rows:
            total = item["total"]
            mistakes = item["mistakes"]
            percent = mistakes / total if total else 0
            users = ", ".join(sorted(item["users"]))

            ws.cell(row=row, column=1, value=item["question_order"])
            ws.cell(row=row, column=2, value=get_test_type_name(item["test_type"]))
            ws.cell(row=row, column=3, value=mistakes)
            ws.cell(row=row, column=4, value=total)
            ws.cell(row=row, column=5, value=percent)
            ws.cell(row=row, column=6, value=users)

            fill = summary_warning_fill if mistakes else summary_ok_fill
            for col in range(1, 7):
                cell = ws.cell(row=row, column=col)
                cell.fill = fill
                cell.border = border
                cell.alignment = center_alignment if col < 6 else Alignment(wrap_text=True, vertical="center")
            ws.cell(row=row, column=5).number_format = "0%"
            row += 1

        ws.column_dimensions['A'].width = 14
        ws.column_dimensions['B'].width = 14
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 14
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 50

        wb.save(filename)

        document = FSInputFile(filename)

        await message.answer_document(
            document=document,
            caption=f"📊 Отчет по тесту №{test.id}",
        )


# =========================
# TEXT HANDLER
# =========================

@dp.message()
async def text_handler(
    message: Message,
    state: FSMContext,
):

    current_state = await state.get_state()

    if not current_state:
        return

    if current_state == TestStates.passing_test.state:
        if not message.text:
            return

        if not message.text.isdigit():
            await message.answer(
                "Введите ответ числом"
            )
            return

        data = await state.get_data()
        question_id = data.get("current_question_id")

        if not question_id:
            await message.answer(
                "❌ Не удалось определить текущий вопрос. Нажмите «Пройти тест» еще раз"
            )
            return

        answer = int(message.text)

        await state.update_data(
            pending_question_id=question_id,
            pending_answer=answer,
        )

        await message.delete()

        kb = InlineKeyboardBuilder()
        kb.button(text="✅ Подтвердить", callback_data="confirm_test_answer:yes")
        kb.button(text="❌ Изменить", callback_data="confirm_test_answer:no")
        kb.adjust(2)

        await message.answer(
            f"Ваш ответ: <b>{answer}</b>\n\nПодтвердить?",
            reply_markup=kb.as_markup(),
            parse_mode=ParseMode.HTML,
        )

        return

    # ДОБАВЛЕНИЕ ВОПРОСА
    if current_state == QuestionAddStates.waiting_answer.state:

        if not message.text:
            return

        if not message.text.isdigit():
            data = await state.get_data()
            if data.get("instruction_message_id"):
                try:
                    await message.bot.delete_message(
                        message.chat.id,
                        data["instruction_message_id"]
                    )
                except:
                    pass

            msg = await message.answer("Введите число")
            await state.update_data(
                instruction_message_id=msg.message_id
            )
            return

        answer = int(message.text)

        data = await state.get_data()

        async with async_session() as session:
            if await is_test_approved(session, data["test_id"]):
                await state.clear()
                await message.answer(
                    "🔐 Тест уже опубликован, добавление вопросов закрыто"
                )
                return
        
        # Сохраняем ответ в state и переходим на подтверждение
        await state.update_data(correct_answer=answer)
        await state.set_state(QuestionAddStates.confirming_question)
        
        await message.delete()
        
        # Удаляем инструкционное сообщение
        if data.get("instruction_message_id"):
            try:
                await message.bot.delete_message(
                    message.chat.id,
                    data["instruction_message_id"]
                )
            except:
                pass
        
        # Показываем превью вопроса с фото и ответом
        kb = InlineKeyboardBuilder()
        kb.button(text="✅ Подтвердить", callback_data="confirm_question:yes")
        kb.button(text="❌ Переснять", callback_data="confirm_question:no")
        kb.adjust(2)
        
        await message.answer_photo(
            photo=data["photo_id"],
            caption=f"📋 Правильный ответ: <b>{answer}</b>\n\n"
                   f"Всё правильно?",
            reply_markup=kb.as_markup(),
            parse_mode=ParseMode.HTML
        )

        return

    # РЕДАКТИРОВАНИЕ ОТВЕТА
    if current_state == AdminEditStates.editing_answer.state:

        if not message.text:
            return

        if not message.text.isdigit():
            data = await state.get_data()
            if data.get("instruction_message_id"):
                try:
                    await message.bot.delete_message(
                        message.chat.id,
                        data["instruction_message_id"]
                    )
                except:
                    pass

            msg = await message.answer("Введите число")
            await state.update_data(
                instruction_message_id=msg.message_id
            )
            return

        answer = int(message.text)

        data = await state.get_data()

        async with async_session() as session:

            result = await session.execute(
                select(Question, Test)
                .join(Test, Test.id == Question.test_id)
                .where(Question.id == data["edit_question_id"])
            )

            row = result.first()

            if not row:
                await state.clear()
                await message.answer("❌ Вопрос не найден")
                return

            q, test = row

            if test.approved:
                await state.clear()
                await message.answer(
                    "🔐 Тест уже опубликован, редактирование закрыто"
                )
                return

            q.correct_answer = answer

            await session.commit()

        await state.clear()

        await message.delete()

        # Удаляем инструкционное сообщение
        if data.get("instruction_message_id"):
            try:
                await message.bot.delete_message(
                    message.chat.id,
                    data["instruction_message_id"]
                )
            except:
                pass

        await message.answer(
            "✅ Ответ обновлён"
        )

        return


# =========================
# MAIN
# =========================

async def setup_bot_commands():
    commands = [
        BotCommand(command="start", description="Начать"),
    ]

    for attempt in range(1, 4):
        try:
            await bot.set_my_commands(commands)
            return
        except Exception as e:
            logger.warning(
                "⚠️ Failed to set bot commands "
                f"(attempt {attempt}/3): {e}"
            )
            await asyncio.sleep(attempt)

    logger.warning(
        "⚠️ Bot commands were not set, continuing startup"
    )


async def main():

    await create_db()
    await setup_bot_commands()

    retry_delay = 5

    while True:
        try:
            logger.info("🚀 Bot started, polling messages...")
            await dp.start_polling(bot)
            break
        except TelegramServerError as e:
            logger.warning(
                "⚠️ Telegram server error during polling startup: "
                f"{e}. Retrying in {retry_delay} seconds..."
            )
            await asyncio.sleep(retry_delay)
            retry_delay = min(retry_delay * 2, 60)


if __name__ == "__main__":
    asyncio.run(main())
